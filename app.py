#!/usr/bin/env python3
"""NFS-e Cloud v6 — com autenticação, deploy-ready"""
import os,sys,base64,gzip,re,csv,io,zipfile,shutil,tempfile,threading,uuid,json
from pathlib import Path
from datetime import datetime,date,timedelta
from xml.etree import ElementTree as ET
from functools import wraps
from flask import Flask,request,jsonify,send_file,Response,send_from_directory

app=Flask(__name__,static_folder="static",static_url_path="/static")
app.config["SECRET_KEY"]=os.environ.get("SECRET_KEY",os.urandom(32).hex())

WORK_DIR=Path(tempfile.gettempdir())/"nfse_work"; WORK_DIR.mkdir(exist_ok=True)
jobs:dict={}
URL_ADN={"producao":"https://adn.nfse.gov.br/contribuintes/DFe","homologacao":"https://adn.producaorestrita.nfse.gov.br/contribuintes/DFe"}
MESES=["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

# ── Auth ──────────────────────────────────────────────────────────────────────
def get_users():
    raw=os.environ.get("USERS","admin:nfse2026,contador:nfse2026")
    u={}
    for p in raw.split(","):
        if ":" in p:
            k,v=p.strip().split(":",1); u[k.strip()]=v.strip()
    return u

def gerar_token(username):
    try:
        import jwt
        return jwt.encode({"sub":username,"exp":datetime.utcnow()+timedelta(hours=12)},app.config["SECRET_KEY"],algorithm="HS256")
    except ImportError:
        import hmac,hashlib,time
        exp=int(time.time())+43200; data=f"{username}:{exp}"
        sig=hmac.new(app.config["SECRET_KEY"].encode(),data.encode(),hashlib.sha256).hexdigest()
        return base64.b64encode(f"{data}:{sig}".encode()).decode()

def verificar_token(token):
    if not token: return None
    try:
        import jwt
        p=jwt.decode(token,app.config["SECRET_KEY"],algorithms=["HS256"]); return p.get("sub")
    except: pass
    try:
        import hmac,hashlib,time
        decoded=base64.b64decode(token.encode()).decode()
        *data_parts,sig=decoded.split(":"); data=":".join(data_parts)
        exp_str=data_parts[-1]; username=":".join(data_parts[:-1])
        expected=hmac.new(app.config["SECRET_KEY"].encode(),data.encode(),hashlib.sha256).hexdigest()
        if hmac.compare_digest(sig,expected) and int(time.time())<int(exp_str): return username
    except: pass
    return None

def requer_login(f):
    @wraps(f)
    def dec(*a,**kw):
        tok=request.headers.get("Authorization","")
        if tok.startswith("Bearer "): tok=tok[7:]
        if not tok: tok=request.args.get("token","")
        if not verificar_token(tok): return jsonify(error="Não autorizado"),401
        return f(*a,**kw)
    return dec

# ── NFS-e logic (igual ao proxy local) ────────────────────────────────────────
def ler_cnpjs(pfx_path,senha):
    try:
        from cryptography.hazmat.primitives.serialization import pkcs12
        import cryptography.x509 as cx
        with open(pfx_path,"rb") as f: raw=f.read()
        _,cert,_=pkcs12.load_key_and_certificates(raw,senha.encode() if senha else None)
        found=set()
        for attr in cert.subject:
            d=''.join(filter(str.isdigit,attr.value))
            if len(d)==14 and d!="00000000000000": found.add(d)
        try:
            san=cert.extensions.get_extension_for_class(cx.SubjectAlternativeName)
            for name in san.value:
                v=getattr(name,'value',b'')
                if isinstance(v,bytes): v=v.decode('ascii',errors='ignore')
                d=''.join(filter(str.isdigit,str(v)))
                if len(d)==14 and d!="00000000000000": found.add(d)
        except: pass
        return sorted(found)
    except: return []

def unzip(b64): return gzip.decompress(base64.b64decode(b64)).decode("utf-8")
def ft(root,*tags):
    for tag in tags:
        for el in root.iter():
            if el.tag.split("}")[-1]==tag and el.text: return el.text.strip()
    return ""
def dt_str(s): m=re.search(r'(\d{4}-\d{2}-\d{2})',s or ""); return m.group(1) if m else ""

def classificar(xml_str,cnpj):
    d=dict(tipo="",numero="",data_emissao="",competencia="",valor="",prestador_cnpj="",prestador_nome="",
           tomador_cnpj="",tomador_nome="",municipio="",chave="",chave_cancelada="",motivo="",
           aliquota="",iss="",deducoes="",base_calculo="",codigo_servico="",descricao_servico="")
    try:
        root=ET.fromstring(xml_str); raiz=root.tag.split("}")[-1]
        if raiz=="evento":
            for el in root.iter():
                if el.tag.split("}")[-1]=="infEvento": d["chave"]=el.get("Id",""); break
            cnpj_autor=""; tp_evento=""
            for el in root.iter():
                t=el.tag.split("}")[-1]
                if t=="CNPJ" and not cnpj_autor and el.text: cnpj_autor=''.join(filter(str.isdigit,el.text))
                if t=="chNFSe" and not d["chave_cancelada"] and el.text: d["chave_cancelada"]=el.text.strip()
                if t=="tpEvento" and el.text: tp_evento=el.text.strip()
                if t in("dhEvento","dhProc") and not d["data_emissao"] and el.text: d["data_emissao"]=dt_str(el.text)
                if t in("xJust","xMotivo","descEvento","descricaoEvento") and not d["motivo"] and el.text: d["motivo"]=el.text.strip()
            CANCEL_CODES={"110110","110111","110112"}
            CANCEL_WORDS={"cancel","cancelar","cancelamento","nao prestado","não prestado","serviço não","servico nao"}
            ml=(d["motivo"] or "").lower(); tl=tp_evento.lower()
            is_cancel=(tp_evento in CANCEL_CODES or any(w in ml for w in CANCEL_WORDS) or "cancel" in tl)
            if tp_evento=="110111": d["tipo"]="substituidas"
            elif is_cancel: d["tipo"]="canceladas"
            else: d["tipo"]="eventos"
            d["prestador_cnpj"]=cnpj_autor or cnpj; return d
        for el in root.iter():
            if el.tag.split("}")[-1]=="infNFSe": d["chave"]=el.get("Id",""); break
        d["numero"]=ft(root,"nNFSe"); d["municipio"]=ft(root,"xLocEmi","xLocPrestacao","xMun")
        d["data_emissao"]=dt_str(ft(root,"dCompet") or ft(root,"dhEmi") or ft(root,"dhProc")); d["competencia"]=d["data_emissao"][:7]
        d["valor"]=ft(root,"vServ") or ft(root,"vLiq") or ft(root,"vBC")
        d["iss"]=ft(root,"vISSQN","vISS"); d["deducoes"]=ft(root,"vDeducao","vDeducoes"); d["base_calculo"]=ft(root,"vBC")
        d["aliquota"]=ft(root,"pAliqAplic","aliqISS","aliq"); d["codigo_servico"]=ft(root,"cServ","cLC116","cTribMun")
        d["descricao_servico"]=ft(root,"xTribNac","xTribMun","xNBS","xDescServ")
        ec=en=pc=tc=tn=""
        for el in root.iter():
            t=el.tag.split("}")[-1]
            if t=="emit":
                for s in el:
                    st=s.tag.split("}")[-1]
                    if st=="CNPJ" and s.text: ec=''.join(filter(str.isdigit,s.text))
                    if st=="xNome" and s.text: en=s.text.strip()
            elif t=="prest":
                for s in el:
                    if s.tag.split("}")[-1]=="CNPJ" and s.text: pc=''.join(filter(str.isdigit,s.text)); break
            elif t=="toma":
                for s in el:
                    st=s.tag.split("}")[-1]
                    if st=="CNPJ" and s.text: tc=''.join(filter(str.isdigit,s.text))
                    if st=="xNome" and s.text: tn=s.text.strip()
        pr=pc or ec; d["prestador_cnpj"]=pr; d["prestador_nome"]=en; d["tomador_cnpj"]=tc; d["tomador_nome"]=tn
        if pr==cnpj: d["tipo"]="emitidas"
        elif tc==cnpj: d["tipo"]="recebidas"
        else:
            ck=d["chave"][9:23] if len(d["chave"])>=23 else ""
            if ck==cnpj: d["tipo"]="emitidas"
    except: pass
    return d

def worker(jid,pfx,senha,cnpjs,ambiente,mes,ano):
    job=jobs[jid]
    def L(tag,msg): job["logs"].append({"ts":datetime.now().strftime("%H:%M:%S"),"tag":tag,"msg":msg})
    try: _run(jid,pfx,senha,cnpjs,ambiente,mes,ano,L)
    finally:
        try:
            if pfx and os.path.exists(pfx): os.unlink(pfx)
        except: pass
        for d in job.get("docs",[]): d.pop("_xml",None)

def _run(jid,pfx,senha,cnpjs,ambiente,mes,ano,L):
    from requests_pkcs12 import get as pkcs12_get
    import time
    job=jobs[jid]
    dt_ini=date(ano,mes,1); dias=[31,28,31,30,31,30,31,31,30,31,30,31][mes-1]
    if mes==2 and(ano%4==0 and(ano%100!=0 or ano%400==0)): dias=29
    dt_fim=date(ano,mes,dias)
    L("info",f"Período: {dt_ini.strftime('%d/%m/%Y')} → {dt_fim.strftime('%d/%m/%Y')}")
    L("info",f"CNPJs: {', '.join(cnpjs)}")
    base=URL_ADN[ambiente]; pasta=WORK_DIR/jid; pasta.mkdir(exist_ok=True)
    stats=dict(analisados=0,no_periodo=0,emitidas=0,recebidas=0,canceladas=0,substituidas=0,eventos=0); docs=[]
    for cnpj in cnpjs:
        L("info",f"▶ CNPJ {cnpj}"); nsu=0; lote=1; parou=False
        while not parou:
            if lote>1: time.sleep(0.1)
            job["nsu_atual"]=nsu; job["lote"]=lote; r=None
            L("info",f"  [{cnpj}] NSU {nsu} lote {lote}...")
            for t in range(1,4):
                try:
                    r=pkcs12_get(f"{base}/{nsu}",params={"cnpjConsulta":cnpj},headers={"Accept":"application/json"},pkcs12_filename=pfx,pkcs12_password=senha,timeout=60)
                    if r.status_code==429:
                        w=20*t; L("warn",f"429 — aguardando {w}s..."); time.sleep(w); r=None; continue
                    break
                except Exception as e:
                    L("warn",f"Tentativa {t}/3 falhou: {e}")
                    if t<3: time.sleep(3*t)
            if r is None: L("erro",f"Falha ({cnpj})."); break
            if r.status_code in(204,404): L("ok",f"Fim ({cnpj})."); break
            if r.status_code in(401,403): L("erro",f"Acesso negado {r.status_code}."); break
            if r.status_code==400:
                # E2243 = CNPJ não autorizado para este certificado — pula para o próximo
                try:
                    erros=r.json().get("Erros",[])
                    codigos=[e.get("Codigo","") for e in erros]
                    if "E2243" in codigos:
                        L("warn",f"  CNPJ {cnpj} não autorizado neste certificado (E2243) — ignorado.")
                        break
                except: pass
                L("erro",f"HTTP 400: {r.text[:150]}"); break
            if r.status_code!=200: L("erro",f"HTTP {r.status_code}: {r.text[:120]}"); break
            try: payload=r.json()
            except: L("erro","Resposta inválida."); break
            if payload.get("StatusProcessamento")=="SEM_DOCUMENTOS": L("ok","Sem documentos."); break
            lote_docs=payload.get("LoteDFe",[])
            if not lote_docs: L("ok","LoteDFe vazio."); break
            L("info",f"  NSU {nsu} → {len(lote_docs)} doc(s)"); stats["analisados"]+=len(lote_docs); ultimo=nsu
            for raw in lote_docs:
                nsu_doc=int(raw.get("NSU",nsu)); b64=raw.get("ArquivoXml",""); chave_ac=raw.get("ChaveAcesso","")
                ultimo=max(ultimo,nsu_doc)
                if not b64: continue
                try: xml_str=unzip(b64)
                except: continue
                m=classificar(xml_str,cnpj); m["nsu"]=nsu_doc; m["chave"]=m["chave"] or chave_ac
                if not m["tipo"]:
                    ck=chave_ac[9:23] if len(chave_ac)>=23 else ""
                    m["tipo"]="emitidas" if ck==cnpj else "recebidas"
                dt=None
                if m["data_emissao"]:
                    try: dt=datetime.strptime(m["data_emissao"][:10],"%Y-%m-%d").date()
                    except: pass
                if dt and dt<dt_ini: continue
                if dt and dt>dt_fim: parou=True; continue
                stats["no_periodo"]+=1; stats[m["tipo"]]=stats.get(m["tipo"],0)+1
                d2=pasta/m["tipo"]; d2.mkdir(exist_ok=True)
                nome=f"{m['chave'] or f'nsu_{nsu_doc}'}.xml"; (d2/nome).write_text(xml_str,encoding="utf-8")
                m["arquivo"]=str(d2/nome); m["_xml"]=xml_str; docs.append(m); job["docs"]=docs; job["stats"]=stats
                if m["tipo"]=="emitidas": emp=m.get("tomador_nome","")
                elif m["tipo"]=="recebidas": emp=m.get("prestador_nome","")
                else: emp=m.get("motivo","") or m.get("chave_cancelada","")[:20]
                L("ok",f"  [{m['tipo']}] NSU {nsu_doc} | {m['data_emissao']} | {emp[:45]}")
            nsu=ultimo+1; lote+=1; job["pct"]=min(90,job.get("pct",0)+1)
        if parou: L("ok",f"  Período concluído para {cnpj}.")
    job.update(pct=100,status="done",fim=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),stats=stats,docs=docs)
    L("ok","─"*52)
    L("ok",f"Analisados: {stats['analisados']} | No período: {stats['no_periodo']}")
    L("ok",f"Emitidas: {stats['emitidas']} | Recebidas: {stats['recebidas']}")
    L("ok",f"Canceladas: {stats['canceladas']} | Substituídas: {stats['substituidas']} | Eventos: {stats['eventos']}")

# ── XLSX e HTML (mesmo do proxy local) ────────────────────────────────────────
def gerar_xlsx(job):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    docs=job.get("docs",[]); stats=job.get("stats",{}); cnpjs=job.get("cnpjs",[])
    mes=job.get("mes",1); ano=job.get("ano",datetime.now().year)
    periodo=f"{MESES[mes-1]}/{ano}"; agora=job.get("fim",datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    CT="1E3A5F";CE="DBEAFE";CR="F3E8FF";CC="FEE2E2";CS="FEF3C7";CN="F3F4F6";CH="1D4ED8";CTO="F0F4FF";WH="FFFFFF"
    wb=Workbook()
    def thin(): s=Side(style='thin',color="D1D5DB"); return Border(left=s,right=s,top=s,bottom=s)
    def hdr(ws,r,c,v,bg=CH,fg=WH,sz=10):
        x=ws.cell(row=r,column=c,value=v); x.font=Font(bold=True,color=fg,size=sz,name="Arial")
        x.fill=PatternFill("solid",start_color=bg); x.alignment=Alignment(horizontal="center",vertical="center"); x.border=thin(); return x
    def cell(ws,r,c,v,bold=False,bg=None,nf=None,al="left",sz=9):
        x=ws.cell(row=r,column=c,value=v); x.font=Font(bold=bold,name="Arial",size=sz)
        if bg: x.fill=PatternFill("solid",start_color=bg)
        x.alignment=Alignment(horizontal=al,vertical="center"); x.border=thin()
        if nf: x.number_format=nf; return x
    def moeda(v):
        try: return float((v or "0").replace(",","."))
        except: return 0.0
    def card(ws,row,col,titulo,valor,cbg,cfg):
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
        c=ws.cell(row=row,column=col,value=titulo); c.font=Font(bold=True,size=10,color=cfg,name="Arial")
        c.fill=PatternFill("solid",start_color=cbg); c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[row].height=20
        ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
        c2=ws.cell(row=row+1,column=col,value=valor); c2.font=Font(bold=True,size=20,color=cfg,name="Arial")
        c2.fill=PatternFill("solid",start_color=cbg); c2.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[row+1].height=32
    ws=wb.active; ws.title="📊 Resumo"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1"); c=ws["A1"]; c.value=f"NFS-e — Resumo: {periodo}"
    c.font=Font(bold=True,size=14,color=WH,name="Arial"); c.fill=PatternFill("solid",start_color=CT)
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:H2"); c=ws["A2"]; c.value=f"CNPJs: {', '.join(cnpjs)} | Gerado em: {agora}"
    c.font=Font(size=9,color="6B7280",name="Arial"); c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18
    card(ws,4,1,"📋 TOTAL",stats.get("no_periodo",0),"E0F2FE","0369A1")
    card(ws,4,3,"⬆ EMITIDAS",stats.get("emitidas",0),"DBEAFE","1D4ED8")
    card(ws,4,5,"⬇ RECEBIDAS",stats.get("recebidas",0),"F3E8FF","7C3AED")
    card(ws,4,7,"✕ CANCELADAS",stats.get("canceladas",0),"FEE2E2","B91C1C")
    card(ws,7,3,"↺ SUBSTITUÍDAS",stats.get("substituidas",0),"FEF3C7","92400E")
    card(ws,7,5,"◎ EVENTOS",stats.get("eventos",0),"F3F4F6","4B5563")
    row=10; ws.merge_cells(f"A{row}:H{row}"); c=ws.cell(row=row,column=1,value="VALORES FINANCEIROS")
    c.font=Font(bold=True,size=11,color=WH,name="Arial"); c.fill=PatternFill("solid",start_color="1E3A5F")
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[row].height=22; row+=1
    for i,h in enumerate(["Categoria","Qtd Notas","Valor Total (R$)","ISS Total (R$)","Média por Nota (R$)"],1): hdr(ws,row,i,h,sz=9)
    ws.row_dimensions[row].height=20; row+=1
    ve=vr=vc=ie=ir=0.0; qe=qr=qc=0
    for d in docs:
        v=moeda(d.get("valor","")); i2=moeda(d.get("iss",""))
        if d["tipo"]=="emitidas": ve+=v; ie+=i2; qe+=1
        elif d["tipo"]=="recebidas": vr+=v; ir+=i2; qr+=1
        elif d["tipo"]=="canceladas": vc+=v; qc+=1
    for cat,qtd,val,iss_v,bg in [("⬆ Emitidas",qe,ve,ie,CE),("⬇ Recebidas",qr,vr,ir,CR),("✕ Canceladas",qc,vc,0.0,CC)]:
        media=val/qtd if qtd else 0
        cell(ws,row,1,cat,bg=bg); cell(ws,row,2,qtd,al="center",bg=bg); cell(ws,row,3,val,bg=bg,nf="R$ #,##0.00",al="right")
        cell(ws,row,4,iss_v,bg=bg,nf="R$ #,##0.00",al="right"); cell(ws,row,5,media,bg=bg,nf="R$ #,##0.00",al="right"); row+=1
    cell(ws,row,1,"TOTAL",bold=True,bg=CTO); cell(ws,row,2,qe+qr,bold=True,bg=CTO,al="center")
    cell(ws,row,3,ve+vr,bold=True,bg=CTO,nf="R$ #,##0.00",al="right"); cell(ws,row,4,ie+ir,bold=True,bg=CTO,nf="R$ #,##0.00",al="right")
    cell(ws,row,5,(ve+vr)/(qe+qr) if(qe+qr)else 0,bold=True,bg=CTO,nf="R$ #,##0.00",al="right")
    for col,w in [("A",20),("B",12),("C",18),("D",16),("E",18),("F",12),("G",12),("H",12)]: ws.column_dimensions[col].width=w
    COLS=[("NSU",10,"nsu"),("Número",10,"numero"),("Data Emissão",14,"data_emissao"),("Competência",12,"competencia"),
          ("Prestador CNPJ",18,"prestador_cnpj"),("Prestador Nome",35,"prestador_nome"),("Tomador CNPJ",18,"tomador_cnpj"),
          ("Tomador Nome",35,"tomador_nome"),("Município",20,"municipio"),("Valor Serviço",16,"valor"),("ISS",12,"iss"),
          ("Base Cálculo",14,"base_calculo"),("Alíquota",10,"aliquota"),("Deduções",12,"deducoes"),
          ("Cód. Serviço",14,"codigo_servico"),("Descrição Serviço",30,"descricao_servico"),("Chave Acesso",55,"chave")]
    COLS_C=[("NSU",10,"nsu"),("Data Evento",14,"data_emissao"),("Tipo",14,"tipo"),("Motivo",45,"motivo"),
            ("Chave Cancelada",55,"chave_cancelada"),("Chave Evento",55,"chave"),("CNPJ Autor",18,"prestador_cnpj")]
    def criar_aba(nome,tp,cols,cor):
        fl=[d for d in docs if d.get("tipo")==tp]
        if not fl: return
        ws2=wb.create_sheet(nome); ws2.sheet_view.showGridLines=False; ws2.freeze_panes="A3"
        ws2.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        c=ws2.cell(row=1,column=1,value=f"{nome} — {periodo} | {len(fl)} nota(s)")
        c.font=Font(bold=True,size=12,color=WH,name="Arial"); c.fill=PatternFill("solid",start_color=CT)
        c.alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=26
        for i,(titulo,w,_) in enumerate(cols,1): hdr(ws2,2,i,titulo,bg=cor,sz=9); ws2.column_dimensions[get_column_letter(i)].width=w
        ws2.row_dimensions[2].height=20; vt=0.0
        for r,d in enumerate(fl,3):
            bg=WH if r%2==0 else "F9FAFB"
            for i,(_,_,campo) in enumerate(cols,1):
                val=d.get(campo,""); fmt=None; al="left"
                if campo=="nsu": al="center"
                if campo=="valor" and val:
                    try: val=float(val.replace(",",".")); fmt="R$ #,##0.00"; al="right"; vt+=val
                    except: pass
                if campo in("iss","base_calculo","deducoes") and val:
                    try: val=float(val.replace(",",".")); fmt="R$ #,##0.00"; al="right"
                    except: pass
                cell(ws2,r,i,val,bg=bg,nf=fmt,al=al)
            ws2.row_dimensions[r].height=16
        tr=len(fl)+3; cell(ws2,tr,1,f"TOTAL — {len(fl)} nota(s)",bold=True,bg=CTO)
        for i in range(2,len(cols)+1): cell(ws2,tr,i,"",bg=CTO)
        for i,(_,_,campo) in enumerate(cols,1):
            if campo=="valor":
                ct=ws2.cell(row=tr,column=i,value=vt); ct.font=Font(bold=True,size=10,name="Arial")
                ct.fill=PatternFill("solid",start_color=CTO); ct.number_format="R$ #,##0.00"
                ct.alignment=Alignment(horizontal="right",vertical="center"); ct.border=thin()
        ws2.row_dimensions[tr].height=20
    criar_aba("⬆ Emitidas","emitidas",COLS,"1D4ED8"); criar_aba("⬇ Recebidas","recebidas",COLS,"7C3AED")
    criar_aba("✕ Canceladas","canceladas",COLS_C,"B91C1C"); criar_aba("↺ Substituídas","substituidas",COLS_C,"D97706"); criar_aba("◎ Eventos","eventos",COLS_C,"4B5563")
    wsa=wb.create_sheet("📋 Todas as Notas"); wsa.sheet_view.showGridLines=False; wsa.freeze_panes="A3"
    CA=[("Tipo",12,"tipo")]+COLS
    wsa.merge_cells(f"A1:{get_column_letter(len(CA))}1"); c=wsa.cell(row=1,column=1,value=f"Todas as Notas — {periodo} | {len(docs)} doc(s)")
    c.font=Font(bold=True,size=12,color=WH,name="Arial"); c.fill=PatternFill("solid",start_color=CT)
    c.alignment=Alignment(horizontal="center",vertical="center"); wsa.row_dimensions[1].height=26
    for i,(titulo,w,_) in enumerate(CA,1): hdr(wsa,2,i,titulo,sz=9); wsa.column_dimensions[get_column_letter(i)].width=w
    wsa.row_dimensions[2].height=20
    CMAP={"emitidas":CE,"recebidas":CR,"canceladas":CC,"substituidas":CS,"eventos":CN}
    for r,d in enumerate(docs,3):
        bg=CMAP.get(d.get("tipo",""),WH)
        for i,(_,_,campo) in enumerate(CA,1):
            val=d.get(campo,""); fmt=None; al="left"
            if campo=="nsu": al="center"
            if campo=="valor" and val:
                try: val=float(val.replace(",",".")); fmt="R$ #,##0.00"; al="right"
                except: pass
            if campo in("iss","base_calculo","deducoes") and val:
                try: val=float(val.replace(",",".")); fmt="R$ #,##0.00"; al="right"
                except: pass
            cell(wsa,r,i,val,bg=bg,nf=fmt,al=al)
        wsa.row_dimensions[r].height=16
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── Rotas API ─────────────────────────────────────────────────────────────────
@app.route("/api/login",methods=["POST"])
def login():
    data=request.get_json() or {}
    u=data.get("username","").strip(); p=data.get("password","").strip()
    users=get_users()
    if u in users and users[u]==p: return jsonify(token=gerar_token(u),username=u)
    return jsonify(error="Usuário ou senha incorretos"),401

@app.route("/api/ping")
@requer_login
def ping(): return jsonify(ok=True,version="6.0")

@app.route("/api/health")
def health(): return jsonify(status="ok")

@app.route("/api/check_cert",methods=["POST"])
@requer_login
def check_cert():
    f=request.files.get("cert")
    if not f: return jsonify(error="Sem cert."),400
    senha=request.form.get("senha","")
    suf=Path(f.filename).suffix or ".pfx"
    tmp=tempfile.NamedTemporaryFile(delete=False,suffix=suf); f.save(tmp.name); tmp.close()
    try: return jsonify(cnpjs=ler_cnpjs(tmp.name,senha))
    except: return jsonify(cnpjs=[])
    finally:
        try: os.unlink(tmp.name)
        except: pass

@app.route("/api/start",methods=["POST"])
@requer_login
def start():
    f=request.files.get("cert")
    if not f or not f.filename: return jsonify(error="Envie o certificado."),400
    senha=request.form.get("senha","")
    if not senha: return jsonify(error="Informe a senha."),400
    mes=int(request.form.get("mes",datetime.now().month)); ano=int(request.form.get("ano",datetime.now().year))
    ambiente=request.form.get("ambiente","producao")
    suf=Path(f.filename).suffix or ".pfx"; tmp=tempfile.NamedTemporaryFile(delete=False,suffix=suf); f.save(tmp.name); tmp.close()
    cnpjs=ler_cnpjs(tmp.name,senha)
    if not cnpjs:
        try: os.unlink(tmp.name)
        except: pass
        return jsonify(error="CNPJ não identificado. Verifique a senha."),400
    jid=str(uuid.uuid4())[:8]
    jobs[jid]=dict(id=jid,status="running",cnpjs=cnpjs,mes=mes,ano=ano,ambiente=ambiente,logs=[],docs=[],stats={},
                   pct=0,nsu_atual=0,lote=0,inicio=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),fim="")
    threading.Thread(target=worker,daemon=True,args=(jid,tmp.name,senha,cnpjs,ambiente,mes,ano)).start()
    return jsonify(job_id=jid,cnpjs=cnpjs)

@app.route("/api/status/<jid>")
@requer_login
def status(jid):
    j=jobs.get(jid)
    if not j: return jsonify(error="não encontrado"),404
    docs=[{k:v for k,v in d.items() if k!="_xml"} for d in j.get("docs",[])]
    return jsonify(status=j["status"],pct=j["pct"],nsu=j.get("nsu_atual",0),lote=j.get("lote",0),
                   cnpjs=j.get("cnpjs",[]),stats=j.get("stats",{}),logs=j.get("logs",[])[-400:],docs=docs,
                   inicio=j.get("inicio",""),fim=j.get("fim",""))

@app.route("/api/zip/<jid>")
@requer_login
def get_zip(jid):
    j=jobs.get(jid)
    if not j: return "não encontrado",404
    buf=io.BytesIO(); pasta=WORK_DIR/jid
    with zipfile.ZipFile(buf,"w",zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(pasta.rglob("*.xml")): zf.write(f,f.relative_to(pasta))
    buf.seek(0); mn=MESES[j.get("mes",1)-1][:3]; an=j.get("ano","")
    return send_file(buf,mimetype="application/zip",as_attachment=True,download_name=f"NFS-e_{mn}{an}_{jid}.zip")

@app.route("/api/xlsx/<jid>")
@requer_login
def get_xlsx(jid):
    j=jobs.get(jid)
    if not j: return "não encontrado",404
    buf=gerar_xlsx(j); mn=MESES[j.get("mes",1)-1][:3]; an=j.get("ano","")
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,download_name=f"NFS-e_{mn}{an}_{jid}.xlsx")

@app.route("/api/relatorio/<jid>")
@requer_login
def get_relatorio(jid):
    j=jobs.get(jid)
    if not j: return "não encontrado",404
    mn=MESES[j.get("mes",1)-1]; an=j.get("ano","")
    # HTML do relatório inline
    docs=j.get("docs",[]); stats=j.get("stats",{}); cnpjs=j.get("cnpjs",[]); mes=j.get("mes",1); ano2=j.get("ano",2026)
    periodo=f"{MESES[mes-1]}/{ano2}"; agora=j.get("fim",datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    vt=sum(float(d.get("valor","0").replace(",",".") or 0) for d in docs if d.get("tipo") in("emitidas","recebidas") and d.get("valor"))
    vf=f"R$ {vt:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    cores={"emitidas":("#dbeafe","#1d4ed8","⬆ Emitida"),"recebidas":("#f3e8ff","#7c3aed","⬇ Recebida"),
           "canceladas":("#fee2e2","#b91c1c","✕ Cancelada"),"substituidas":("#fef3c7","#92400e","↺ Substituída"),"eventos":("#f3f4f6","#4b5563","◎ Evento")}
    linhas="".join(f"""<tr><td><span style="background:{cores.get(d['tipo'],('#eee','#333','?'))[0]};color:{cores.get(d['tipo'],('#eee','#333','?'))[1]};padding:2px 8px;border-radius:100px;font-size:11px;font-weight:600">{cores.get(d['tipo'],('#eee','#333',d['tipo']))[2]}</span></td>
      <td class=m>{d.get('numero','—')}</td><td>{d.get('data_emissao','—')}</td>
      <td style="max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{d.get('tomador_nome','') if d['tipo']=='emitidas' else d.get('prestador_nome','—')}</td>
      <td>{d.get('municipio','—')}</td><td class="m r">{d.get('valor','—')}</td>
      <td class="m s">{d.get('chave','')[:26]}{"…" if len(d.get('chave',''))>26 else ""}</td></tr>""" for d in docs)
    html=f"""<!DOCTYPE html><html lang=pt-BR><head><meta charset=UTF-8><title>NFS-e {periodo}</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:-apple-system,"Segoe UI",sans-serif;background:#f4f5f7;color:#1a1d2e;font-size:14px}}
header{{background:#0f172a;color:#fff;padding:1.5rem 2rem;display:flex;align-items:center;justify-content:space-between}}
h1{{font-size:20px;font-weight:700}}h1 em{{color:#38bdf8;font-style:normal}}.meta{{font-size:12px;color:#94a3b8;margin-top:4px}}
.periodo{{font-size:12px;padding:5px 14px;border-radius:100px;font-weight:600;background:rgba(56,189,248,.15);color:#38bdf8;border:1px solid rgba(56,189,248,.3)}}
.w{{max-width:1300px;margin:2rem auto;padding:0 2rem}}.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(135px,1fr));gap:1rem;margin-bottom:2rem}}
.card{{background:#fff;border-radius:10px;padding:1.1rem;border:1px solid #e8eaef}}.cl{{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:#888;margin-bottom:5px}}
.cv{{font-size:24px;font-weight:700}}.gn{{color:#059669}}.bl{{color:#2563eb}}.pu{{color:#7c3aed}}.rd{{color:#dc2626}}.am{{color:#d97706}}.gr{{color:#4b5563}}.vi{{color:#7c3aed;font-size:18px}}
.st{{font-size:11.5px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;color:#888;margin-bottom:.8rem;padding-bottom:6px;border-bottom:1px solid #e8eaef}}
.tw{{background:#fff;border-radius:10px;border:1px solid #e8eaef;overflow:hidden;margin-bottom:2rem}}.ts{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:12.5px}}thead th{{background:#f8f9fb;padding:9px 1rem;text-align:left;font-size:9.5px;text-transform:uppercase;letter-spacing:.09em;color:#888;border-bottom:1px solid #e8eaef;white-space:nowrap}}
tbody tr{{border-bottom:1px solid #f0f1f4}}tbody tr:hover{{background:#f8f9fb}}tbody td{{padding:9px 1rem;vertical-align:middle}}
.m{{font-family:monospace}}.r{{text-align:right}}.s{{font-size:10px;color:#888}}footer{{text-align:center;padding:1.5rem;color:#aaa;font-size:11px}}</style></head><body>
<header><div><h1>NFS-e <em>Lote</em></h1><p class=meta>Gerado em {agora} · CNPJs: {', '.join(cnpjs)}</p></div><span class=periodo>📅 {periodo}</span></header>
<div class=w><div class=cards style="margin-top:1.5rem">
<div class=card><div class=cl>Total</div><div class="cv gn">{stats.get('no_periodo',0)}</div></div>
<div class=card><div class=cl>Emitidas</div><div class="cv bl">{stats.get('emitidas',0)}</div></div>
<div class=card><div class=cl>Recebidas</div><div class="cv pu">{stats.get('recebidas',0)}</div></div>
<div class=card><div class=cl>Canceladas</div><div class="cv rd">{stats.get('canceladas',0)}</div></div>
<div class=card><div class=cl>Valor total</div><div class="cv vi">{vf}</div></div>
</div><div class=st>Notas — {periodo}</div>
<div class=tw><div class=ts><table><thead><tr><th>Tipo</th><th>Número</th><th>Emissão</th><th>Tomador/Prestador</th><th>Município</th><th>Valor R$</th><th>Chave</th></tr></thead>
<tbody>{"".join(linhas) or '<tr><td colspan=7 style="text-align:center;padding:2rem;color:#aaa">Nenhuma nota.</td></tr>'}</tbody></table></div></div></div>
<footer>NFS-e Lote · ADN · {agora}</footer></body></html>"""
    return Response(html,mimetype="text/html",headers={"Content-Disposition":f"attachment; filename=Relatorio_{mn}{an2}_{jid}.html"})

@app.route("/api/limpar/<jid>",methods=["POST"])
@requer_login
def limpar(jid):
    pasta=WORK_DIR/jid
    try:
        if pasta.exists(): shutil.rmtree(pasta)
    except: pass
    jobs.pop(jid,None); return jsonify(ok=True)

# ── Serve o frontend ──────────────────────────────────────────────────────────
@app.route("/",defaults={"path":""})
@app.route("/<path:path>")
def frontend(path):
    static_dir=Path(__file__).parent/"static"
    fp=static_dir/path
    if path and fp.exists(): return send_from_directory(str(static_dir),path)
    return send_from_directory(str(static_dir),"index.html")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    print(f"\n  NFS-e Cloud v6 — porta {port}\n")
    app.run(host="0.0.0.0",port=port,debug=False,threaded=True)
